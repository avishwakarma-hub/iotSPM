from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from utils.review_filters import keep_review_row


def build_review_report(cfg: Dict[str, Any], spm_path: str | Path) -> Path:
    rows = _read_dicts(spm_path)
    rows = [row for row in rows if keep_review_row(row, cfg)]
    out_dir = Path(cfg["paths"]["reports_dir"])
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / f"{Path(spm_path).stem}.review.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "IoT UA Review"
    headers = [
        "Priority", "Hits", "Group Size", "Hardware Type", "Vendor", "Model", "Marketing Name",
        "Candidate Reason", "SPM Status", "SPM Source", "KB/Live Verdict", "Live IoT Matches",
        "KB Match", "KB RefID", "KB SMStat ID", "KB Signature ID",
        "KB Device Type", "KB Pattern", "KB Row Pattern", "KB Dependencies", "KB Match Terms",
        "KB Match Type", "KB Family", "KB Family Size", "Consolidation Note", "Action",
        "Suggested Signature Seed", "User-Agent",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    for idx, row in enumerate(rows, start=1):
        status = row.get("spm_detection_status", "")
        action = _suggest_action(status, row)
        ws.append([
            idx,
            _to_int(row.get("total_group_hits")),
            _to_int(row.get("group_size")),
            row.get("hardware_type", ""),
            row.get("device_vendor", ""),
            row.get("device_model", ""),
            row.get("marketing_name", ""),
            row.get("iot_candidate_reason", ""),
            status,
            row.get("spm_result_source", ""),
            row.get("kb_live_verdict", ""),
            _to_int(row.get("live_iot_match_count")),
            row.get("kb_match", ""),
            row.get("kb_refid", ""),
            row.get("kb_smstat_id", ""),
            row.get("kb_signature_id", ""),
            row.get("kb_device_type", ""),
            row.get("kb_pattern", ""),
            row.get("kb_row_pattern", ""),
            row.get("kb_dependency_patterns", ""),
            row.get("kb_match_terms", ""),
            row.get("kb_match_type", ""),
            row.get("kb_family", ""),
            _to_int(row.get("kb_family_size")),
            _consolidation_note(row),
            action,
            _signature_seed(row.get("user_agent", "")),
            row.get("user_agent", ""),
        ])
    for col in ws.columns:
        max_len = min(max(len(str(cell.value or "")) for cell in col) + 2, 80)
        ws.column_dimensions[col[0].column_letter].width = max_len
    wb.save(output_path)
    return output_path


def focus_report_path(cfg: Dict[str, Any], spm_path: str | Path) -> Path:
    out_dir = Path(cfg["paths"]["reports_dir"])
    return out_dir / f"{Path(spm_path).stem}.focus.xlsx"


def build_focus_report(cfg: Dict[str, Any], spm_path: str | Path) -> Path:
    """Build a separate high-hit model cluster report for signature planning.

    The normal review report is UA-row oriented. This report is intentionally
    model/family oriented so a reviewer can cover the largest Zscaler traffic
    clusters first instead of reviewing hundreds/thousands of near-duplicate UA
    strings that only differ by app/build/version numbers.
    """

    focus_cfg = cfg.get("focus_report", {})
    top_clusters = int(focus_cfg.get("top_clusters", 40))
    max_samples = int(focus_cfg.get("max_sample_uas_per_cluster", 5))
    rows = [row for row in _read_dicts(spm_path) if keep_review_row(row, cfg)]
    actionable_rows = [row for row in rows if row.get("spm_detection_status") in {"not-present", "detected-disabled", "spm-error"}]
    detected_rows = [row for row in rows if row.get("spm_detection_status") not in {"not-present", "detected-disabled", "spm-error"}]
    total_hits = sum(_hits(row) for row in actionable_rows)
    clusters = _cluster_rows(actionable_rows)
    clusters.sort(key=lambda cluster: (cluster["total_hits"], cluster["ua_groups"]), reverse=True)
    detected_clusters = _cluster_rows(detected_rows)
    detected_clusters.sort(key=lambda cluster: (cluster["total_hits"], cluster["ua_groups"]), reverse=True)

    out_dir = Path(cfg["paths"]["reports_dir"])
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = focus_report_path(cfg, spm_path)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Focus Clusters"
    detail_ws = wb.create_sheet("Cluster UA Samples")
    detected_ws = wb.create_sheet("Detected Summary")

    summary_headers = [
        "Priority", "Cluster Key", "Cluster Label", "Total Hits", "Traffic %", "Cumulative %",
        "UA Groups", "Not-present Hits", "Detected-reviewed Hits", "Disabled Hits", "SPM Error Hits",
        "Not-present Groups", "Dominant SPM Status", "Hardware Type", "Vendor", "Model", "Marketing Name",
        "UA Family", "Candidate Reason", "Suggested Action", "Suggested Signature Seed", "Sample UA",
    ]
    detail_headers = [
        "Cluster Priority", "Cluster Label", "Sample Rank", "Hits", "Group Size", "SPM Status",
        "Hardware Type", "Vendor", "Model", "Marketing Name", "UA Family", "Suggested Signature Seed", "User-Agent",
    ]
    detected_headers = [
        "Priority", "Cluster Label", "Total Hits", "UA Groups", "Dominant SPM Status",
        "Hardware Type", "Vendor", "Model", "Marketing Name", "UA Family", "Sample UA",
    ]
    _write_headers(summary_ws, summary_headers, fill="FCE4D6")
    _write_headers(detail_ws, detail_headers, fill="D9EAD3")
    _write_headers(detected_ws, detected_headers, fill="D9EAF7")

    cumulative_hits = 0
    for priority, cluster in enumerate(clusters[:top_clusters], start=1):
        cumulative_hits += int(cluster["total_hits"])
        sample_row = cluster["sample_rows"][0] if cluster["sample_rows"] else {}
        suggested_seed = _cluster_signature_seed(cluster)
        summary_ws.append([
            priority,
            cluster["cluster_key"],
            cluster["label"],
            cluster["total_hits"],
            _pct_number(cluster["total_hits"], total_hits),
            _pct_number(cumulative_hits, total_hits),
            cluster["ua_groups"],
            cluster["status_hits"].get("not-present", 0),
            cluster["status_hits"].get("detected-reviewed", 0),
            cluster["status_hits"].get("detected-disabled", 0),
            cluster["status_hits"].get("spm-error", 0),
            cluster["status_groups"].get("not-present", 0),
            _dominant(cluster["status_hits"]),
            cluster["hardware_type"],
            cluster["vendor"],
            cluster["model"],
            cluster["marketing_name"],
            cluster["ua_family"],
            _dominant(cluster["candidate_reasons"]),
            _cluster_action(cluster),
            suggested_seed,
            sample_row.get("user_agent", ""),
        ])
        for sample_rank, row in enumerate(cluster["sample_rows"][:max_samples], start=1):
            detail_ws.append([
                priority,
                cluster["label"],
                sample_rank,
                _hits(row),
                _to_int(row.get("group_size")),
                row.get("spm_detection_status", ""),
                row.get("hardware_type", ""),
                row.get("device_vendor", ""),
                row.get("device_model", ""),
                row.get("marketing_name", ""),
                _ua_family(row.get("user_agent", "")),
                _signature_seed(row.get("user_agent", "")),
                row.get("user_agent", ""),
            ])

    for priority, cluster in enumerate(detected_clusters[:top_clusters], start=1):
        sample_row = cluster["sample_rows"][0] if cluster["sample_rows"] else {}
        detected_ws.append([
            priority,
            cluster["label"],
            cluster["total_hits"],
            cluster["ua_groups"],
            _dominant(cluster["status_hits"]),
            cluster["hardware_type"],
            cluster["vendor"],
            cluster["model"],
            cluster["marketing_name"],
            cluster["ua_family"],
            sample_row.get("user_agent", ""),
        ])

    for ws in (summary_ws, detail_ws, detected_ws):
        _autosize(ws)
        ws.freeze_panes = "A2"
    wb.save(output_path)
    return output_path


def _suggest_action(status: str, row: Dict[str, Any] | None = None) -> str:
    row = row or {}
    if row.get("kb_live_verdict") == "live-only-likely-review-mode-or-kb-miss":
        return "already-covered-live-review-kb-miss"
    if row.get("kb_match") == "yes" and status == "not-present":
        return "review-existing-kb-match"
    if status in {"not-present", "detected-disabled"}:
        return "review-add-signature"
    if status == "spm-error":
        return "retry-spm/manual-check"
    return "already-covered"


def _cluster_rows(rows: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    clusters: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        cluster_key, attrs = _cluster_identity(row)
        cluster = clusters.setdefault(
            cluster_key,
            {
                "cluster_key": cluster_key,
                "label": _cluster_label(attrs),
                "hardware_type": attrs["hardware_type"],
                "vendor": attrs["vendor"],
                "model": attrs["model"],
                "marketing_name": attrs["marketing_name"],
                "ua_family": attrs["ua_family"],
                "total_hits": 0,
                "ua_groups": 0,
                "status_hits": defaultdict(int),
                "status_groups": Counter(),
                "candidate_reasons": Counter(),
                "sample_rows": [],
            },
        )
        hits = _hits(row)
        status = row.get("spm_detection_status") or "unknown"
        cluster["total_hits"] += hits
        cluster["ua_groups"] += 1
        cluster["status_hits"][status] += hits
        cluster["status_groups"][status] += 1
        if row.get("iot_candidate_reason"):
            cluster["candidate_reasons"][row.get("iot_candidate_reason", "")] += 1
        cluster["sample_rows"].append(row)

    for cluster in clusters.values():
        cluster["sample_rows"].sort(key=_hits, reverse=True)
    return list(clusters.values())


def _cluster_identity(row: Dict[str, Any]) -> Tuple[str, Dict[str, str]]:
    hardware_type = _clean_token(row.get("hardware_type")) or "unknown-hardware"
    vendor = _clean_token(row.get("device_vendor")) or "unknown-vendor"
    model = _clean_token(row.get("device_model"))
    marketing_name = _clean_token(row.get("marketing_name"))
    ua_family = _ua_family(row.get("user_agent", ""))

    # Prefer DeviceAtlas model/marketing identity. Fall back to UA family only
    # when DeviceAtlas cannot identify a model, which keeps app-only IoT UAs
    # grouped without merging unrelated known device models.
    model_or_marketing = model or marketing_name
    identity_model = model_or_marketing or ua_family or "unknown-model"
    key = "|".join([hardware_type, vendor, identity_model, ua_family or "unknown-ua-family"]).casefold()
    return key, {
        "hardware_type": hardware_type,
        "vendor": vendor,
        "model": model,
        "marketing_name": marketing_name,
        "ua_family": ua_family,
    }


def _cluster_label(attrs: Dict[str, str]) -> str:
    device_parts = [attrs.get("vendor", ""), attrs.get("model", "") or attrs.get("marketing_name", "")]
    label = " ".join(part for part in device_parts if part and not part.startswith("unknown-"))
    if not label:
        label = attrs.get("ua_family") or attrs.get("hardware_type") or "Unknown cluster"
    if attrs.get("ua_family") and attrs["ua_family"] not in label:
        label = f"{label} / {attrs['ua_family']}"
    return label


def _cluster_action(cluster: Dict[str, Any]) -> str:
    if cluster["status_hits"].get("not-present", 0) > 0:
        return "focus-review-add/broaden-signature"
    if cluster["status_hits"].get("detected-disabled", 0) > 0:
        return "focus-review-disabled-signature"
    if cluster["status_hits"].get("spm-error", 0) > 0:
        return "retry-spm-before-signature-decision"
    return "monitor-existing-coverage"


def _cluster_signature_seed(cluster: Dict[str, Any]) -> str:
    for row in cluster.get("sample_rows", []):
        if row.get("spm_detection_status") in {"not-present", "detected-disabled"}:
            return _signature_seed(row.get("user_agent", ""))
    if cluster.get("sample_rows"):
        return _signature_seed(cluster["sample_rows"][0].get("user_agent", ""))
    return ""


def _signature_seed(ua: str) -> str:
    # Conservative-but-useful seed: remove volatile version/build tokens while
    # preserving product/model tokens. Reviewer can still choose exact UA if
    # needed, but this makes the focus report immediately useful for clustering.
    seed = str(ua or "").strip()
    seed = re.sub(r"Build/[A-Za-z0-9._+-]+", "Build/*", seed)
    seed = re.sub(r"/\d+(?:\.\d+){1,5}(?:[A-Za-z0-9._+-]*)", "/*", seed)
    seed = re.sub(r"\b\d{6,}\b", "*", seed)
    return seed


def _ua_family(ua: str) -> str:
    first = str(ua or "").strip().split(" ", 1)[0].strip()
    first = first.split("(", 1)[0].strip()
    first = re.sub(r"/[0-9][A-Za-z0-9._+-]*$", "", first)
    first = re.sub(r"/[0-9.]+$", "", first)
    return first or "unknown_app"


def _consolidation_note(row: Dict[str, Any]) -> str:
    if row.get("kb_match") != "yes":
        return ""
    family = str(row.get("kb_family") or "").strip()
    family_size = _to_int(row.get("kb_family_size"))
    pattern = str(row.get("kb_pattern") or "").strip()
    device_type = str(row.get("kb_device_type") or "").strip()
    if family and family_size > 1:
        return (
            f"Fits existing {device_type} family '{family}' with {family_size} SPM patterns; "
            "review whether to consolidate/broaden the family instead of adding a narrow one-off signature."
        )
    return f"Matches existing SPM pattern '{pattern}' ({device_type}); verify why live SPM status differs if marked not-present."


def _read_dicts(path: str | Path) -> List[Dict[str, Any]]:
    with Path(path).open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _to_int(value: Any) -> int:
    try:
        return int(float(value or 0))
    except Exception:
        return 0


def _hits(row: Dict[str, Any]) -> int:
    return _to_int(row.get("total_group_hits") or row.get("hit_count"))


def _clean_token(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    return text


def _dominant(counter: Dict[str, int] | Counter[str]) -> str:
    if not counter:
        return ""
    return max(counter.items(), key=lambda item: item[1])[0]


def _pct_number(value: Any, total: Any) -> float:
    total_int = _to_int(total)
    if total_int <= 0:
        return 0.0
    return round((_to_int(value) / total_int) * 100, 2)


def _write_headers(ws, headers: List[str], *, fill: str) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = min(max(len(str(cell.value or "")) for cell in col) + 2, 90)
        ws.column_dimensions[col[0].column_letter].width = max_len
